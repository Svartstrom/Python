# coding=utf-8

from collections import deque
import glob
import sqlite3
import os
from openpyxl import load_workbook

def Table_exists(f):
    def wraper(*args, **kwargs):
        #print(args)
        #print(kwargs)
        if "symbol" not in kwargs or "db" not in kwargs:
            raise "CallingError"
        #print("CREATE TABLE ? ('date' DATE PRIMARY KEY UNIQUE,'list' TEXT, 'sector' TEXT, 'branch' TEXT, 'direktavkastning' INTEGER, 'pe' INTEGER, 'ps' INTEGER,'pb'INTEGER,'openprice' INTEGER,'highprice'INTEGER,'lowprice'INTEGER, 'closeprice' INTEGER,'volume'INTEGER)",(kwargs['symbol'],))
        try:
            #resp = db.execute("SELECT id FROM transactions WHERE id=?;",(ii,))
            #kwargs['db'].execute("SELECT * FROM ACAD WHERE id=1;")
            SQLinp = "SELECT * FROM " + kwargs['symbol'] + ";"
            kwargs['db'].execute(SQLinp)
        except:
            SQLinp = "CREATE TABLE "+kwargs['symbol']+" ('date' DATE PRIMARY KEY UNIQUE,'list' TEXT, 'sector' TEXT, 'branch' TEXT, 'direktavkastning' INTEGER, 'pe' INTEGER, 'ps' INTEGER,'pb'INTEGER,'openprice' INTEGER,'highprice'INTEGER,'lowprice'INTEGER, 'closeprice' INTEGER,'volume'INTEGER)"
            print(kwargs['symbol'])
            print(SQLinp)
            kwargs['db'].execute(SQLinp)
        return f(*args, **kwargs)
    return wraper
								   
@Table_exists
def insertBD(input_date, trade_list, sector, branch, dir_avk, pe,ps,pb,close_price,symbol,db):
    SQLinp = "SELECT * FROM " + symbol + " WHERE date=?;"
    temp = db.execute(SQLinp,(input_date,))
    if temp:
        SQLinp = "UPDATE " + symbol + " SET list=?, sector=?, branch=?, direktavkastning=?, pe=?, ps=?, pb=?,closeprice=? WHERE date=?"
        db.execute(SQLinp,(trade_list,sector,branch,dir_avk,pe,ps,pb,close_price,input_date,))
    else:
        SQLinp = "INSERT INTO " + symbol + " (datum, list,  sector, branch, direktavkastning, pe, ps, pb,closeprice) VALUES (?,?,?,?,?,?,?,?,?,?);"
        db.execute(SQLinp,(input_date, trade_list, sector, branch, dir_avk, pe, ps, pb, close_price))

@Table_exists
def insertTrades(input_date,open_price,high_price,low_price,close_price,volume,symbol,db):
    SQLinp = "SELECT * FROM " + symbol + " WHERE date=?;"
    temp = db.execute(SQLinp,(input_date,))
    if temp:
        SQLinp = "UPDATE " + symbol + " SET openprice=?, highprice=?,lowprice=?,closeprice=?,volume=? WHERE date=?;"
        db.execute(SQLinp,(open_price,high_price,low_price,close_price,volume,input_date))
    else:
        SQLinp = "INSERT INTO ? (datum,openprice,highprice,lowprice,closeprice,volume) VALUES (?,?,?,?,?,?)"
        db.execute(SQLinp,(input_date,open_price,high_price,low_price,close_price,volume))

# SELECT count(*) FROM sqlite_master WHERE type='table' AND name='table_name';
def createDatabase(DBname):
    db_ = sqlite3.connect(DBname)
    db  = db_.cursor()
    db.execute("""CREATE TABLE 'companies'('symbol' TEXT PRIMARY KEY UNIQUE,'datum' DATE, 'list' TEXT, 'sector' TEXT, 'branch' TEXT, 'direktavkastning' INTEGER, 'pe' INTEGER, 'ps' INTEGER, 'pb' INTEGER);""")
    db.execute("""CREATE TABLE 'transactions'(	'id' INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE, 'symbol' TEXT, 'datum' DATE, 'openprice' INTEGER,'highprice'INTEGER,'lowprice'INTEGER, 'closeprice' INTEGER,'volume'INTEGER );""")
    return db_

def formatName(name):
    name = name.split('/')[-1]
    name = name.replace("_","-")
    name = name.replace(".","-")
    name = name.split('-')
    name[0] = "_"+name[0]
    return name

def readFromXLSX(db,ORGname,ij,max_ant):

    name = formatName(ORGname)
    
    wb = load_workbook(ORGname)
    ws = wb.active
    
    if name[0] == '_Borsdata':
        input_date = str(name[1])+"-"+str(name[2])+"-"+str(name[3])
        #Bolagsnamn	Info	Info	Info	Info	Info	Info		Info	Kursutveck.	Direktav.	P/E	P/S	P/B	Aktiekurs
        #		Land	Lista	Sektor	Bransch	Ticker	Instrument	Rapport	Utveck.  1 år	Senaste		Senaste	Senaste	Senaste	Senaste
        for row in ws.iter_rows(min_row = 3, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            symbol = row[5].value.replace(" ","").replace("-","_").replace(".","_")
            symbol = "_"+symbol
            trade_list = row[2].value
            sector  = row[3].value
            branch  = row[4].value
            dir_avk = row[9].value
            pe = row[10].value
            ps = row[11].value
            pb = row[12].value
            close_price = row[13].value
            resp = insertBD(input_date,trade_list,sector,branch,dir_avk,pe,ps,pb,close_price,symbol=symbol,db=db)
    elif name[0] == 'AAA':
        pass
    else:
        #for row in ws.iter_rows(min_row = 3, min_col=1, max_row=ws.max_row, max_col=ws.max_col):
        for row in ws.iter_rows(min_row = 2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            print(name[0])
            print(row[0].value)
            input_date  = row[0].value.date()
            open_price  = row[1].value
            high_price  = row[2].value
            low_price   = row[3].value
            close_price = row[4].value
            volume = row[5].value
            symbol = name[0]
            insertTrades(input_date,open_price,high_price,low_price,close_price,volume,symbol=symbol,db=db)

def main():
    os.system("rm -f ekonomen.db")
    resp = glob.glob("ekonomen.db")
    if len(resp) < 1:
        db_ = createDatabase("ekonomen.db")
    else:
        db_ = sqlite3.connect("ekonomen.db")
    db = db_.cursor()

    ij = 1

    full_list = glob.glob("Data_test/*.xlsx")
    for ORGname in full_list:
        readFromXLSX(db,ORGname,ij,len(full_list))
        ij+=1
        #os.system("rm -f " + ORGname)
    db_.commit()
    db_.close()
    
if __name__ == "__main__":
    main()
