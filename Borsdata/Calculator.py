import openpyxl
#from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import math
import glob
import datetime
import time
import os

import Swingtrader_constants as C

def printFullMAtoCurrentWB(N,DataDir):

	SHORT_TIMEFRAME  = []
	MEDIUM_TIMEFRAME = []
	LONG_TIMEFRAME 	 = []

	wb = openpyxl.load_workbook(N)
	ws = wb.active

	ws.cell(row = C.NAME_ROW, column = C.MA15_COL ).value  = "MA"+str(C.ShortAmountDays)
	ws.cell(row = C.NAME_ROW, column = C.MA50_COL ).value  = "MA"+str(C.MediumAmountDays)
	ws.cell(row = C.NAME_ROW, column = C.MA200_COL).value = "MA" +str(C.LongAmountDays)

	for i in range(ws.max_row,1,-1):
		Cval = ws.cell(row = i, column = C.Closeprice_COL).value

		SHORT_TIMEFRAME.append(Cval)
		if len(SHORT_TIMEFRAME) == C.ShortAmountDays:
			S = sum(SHORT_TIMEFRAME) / C.ShortAmountDays
			ws.cell(row = i, column = C.MA15_COL).value = S
			del(SHORT_TIMEFRAME[0])
		else:
			ws.cell(row = i, column = C.MA15_COL).value = 'NaN'

		MEDIUM_TIMEFRAME.append(Cval)
		if len(MEDIUM_TIMEFRAME) == C.MediumAmountDays:
			S = sum(MEDIUM_TIMEFRAME) / C.MediumAmountDays
			ws.cell(row = i, column = C.MA50_COL).value = S
			del(MEDIUM_TIMEFRAME[0])
		else:
			ws.cell(row = i, column = C.MA50_COL).value = 'NaN'

		LONG_TIMEFRAME.append(Cval)
		if len(LONG_TIMEFRAME) == C.LongAmountDays:
			S = sum(LONG_TIMEFRAME) / C.LongAmountDays
			ws.cell(row = i, column = C.MA200_COL).value = S
			del(LONG_TIMEFRAME[0])
		else:
			ws.cell(row = i, column = C.MA200_COL).value = 'NaN'
	wb.save(N)

def DeriveAndPrintFullDMAtoCurrentWB(N,DataDir):
	wb = openpyxl.load_workbook(N)
	ws = wb.active
	if ws.cell(row = C.NAME_ROW, column = C.MA15_COL).value != "MA15": return -1

	D_SHORT_TIMEFRAME  = []
	dma50  = []
	dma200 = []	

	ws.cell(row = C.NAME_ROW, column = C.DMA15_COL).value  = "DMA15"
	ws.cell(row = C.NAME_ROW, column = C.DMA50_COL).value  = "DMA50"
	ws.cell(row = C.NAME_ROW, column = C.DMA200_COL).value = "DMA200"

	prevMA15  = ws.cell(row = ws.max_row, column = C.MA15_COL ).value
	prevMA50  = ws.cell(row = ws.max_row, column = C.MA50_COL ).value
	prevMA200 = ws.cell(row = ws.max_row, column = C.MA200_COL).value

	for i in range(ws.max_row,1,-1):
		if ws.cell(row = i, column = C.MA15_COL).value != 'NaN':
			if prevMA15 == 'NaN':
				prevMA15 = ws.cell(row = i, column = C.MA15_COL).value
				ws.cell(row = i, column = C.DMA15_COL).value = 'NaN'
			else:
				Cval = ws.cell(row = i, column = C.MA15_COL).value
				ws.cell(row = i, column = C.DMA15_COL).value = Cval - prevMA15
				prevMA15 = Cval
		else:	
			ws.cell(row = i, column = C.DMA15_COL).value = 'NaN'

		if ws.cell(row = i, column = C.MA50_COL).value != 'NaN':
			if prevMA50 == 'NaN':
				prevMA50 = ws.cell(row = i, column = C.MA50_COL).value
				ws.cell(row = i, column = C.DMA50_COL).value = 'NaN'
			else:
				Cval = ws.cell(row = i, column = C.MA50_COL).value
				ws.cell(row = i, column = C.DMA50_COL).value = Cval - prevMA50
				prevMA50 = Cval
		else:	
			ws.cell(row = i, column = C.DMA50_COL).value = 'NaN'

		if ws.cell(row = i, column = C.MA200_COL).value != 'NaN':
			if prevMA200 == 'NaN':
				prevMA200 = ws.cell(row = i, column = C.MA200_COL).value
				ws.cell(row = i, column = C.DMA200_COL).value = 'NaN'
			else:
				Cval = ws.cell(row = i, column = C.MA200_COL).value
				ws.cell(row = i, column = C.DMA200_COL).value = Cval - prevMA200
				prevMA200 = Cval
		else:	
			ws.cell(row = i, column = C.DMA200_COL).value = 'NaN'
	wb.save(N)

def deleteFullMAandDMA(N):
	wb = openpyxl.load_workbook(N)
	ws = wb.active

	for R in range(1,ws.max_row+1):
		ws.cell(row = R, column = C.MA15_COL  ).value = ""
		ws.cell(row = R, column = C.MA50_COL  ).value = ""
		ws.cell(row = R, column = C.MA200_COL ).value = ""
		ws.cell(row = R, column = C.DMA15_COL ).value = ""
		ws.cell(row = R, column = C.DMA50_COL ).value = ""
		ws.cell(row = R, column = C.DMA200_COL).value = ""
	wb.save(N)