# def importHistory(ws, Rws, D, score):
# def insertRow(Iws,r,ws,A,Ticker_COL,Lista_COL,Kurs_COL,Namn_COL,Datum):
# def importFromDaily():


import openpyxl
#from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import math
import glob
import datetime
import time
import os

import Swingtrader_constants as C

def importWeekHistory(ws, Rws, D, score):
	""" Find the reports for the prev 30 days and insert to Rws """
	count = 1
	C_maxrow = Rws.max_row
	for jj in [2,7]:
		LW = D - datetime.timedelta(days = jj)
		LastWeek = str(LW.year)+'_'+str(LW.month)+'_'+str(LW.day)
		LastWeekR = C.DataDir+'/Reports/'+str(LastWeek)+'_report.xlsx'
		if os.path.isfile(LastWeekR):
			
			LRwb = openpyxl.load_workbook(LastWeekR)
			LRws = LRwb.active
			
			Rws.cell(row = C.NAME_ROW, column = 2+count).value = LastWeek
			
			MR = LRws.max_row
			if MR==0: MR+=1
			
			for l in range(1,MR+1):
				#print(l,LRws.cell(row=l, column=1).value)
				if LRws.cell(row = l, column = C.Date_COL).value == ws.title:
					P = LRws.cell(row = l, column = C.Openprice_COL).value

					Rws.cell(row = C_maxrow, column = 2+count).value = score - P
					break
			count += 1
	
	

def importHistory(ws, Rws, D, score):
	""" Find the reports for the prev 30 days and insert to Rws """
	count = 1
	C_maxrow = Rws.max_row
	for jj in range(1, 30):
		LW = D - datetime.timedelta(days = jj)
		LastWeek = str(LW.year)+'_'+str(LW.month)+'_'+str(LW.day)
		LastWeekR = C.DataDir+'/Reports/'+str(LastWeek)+'_report.xlsx'
		if os.path.isfile(LastWeekR):
			
			LRwb = openpyxl.load_workbook(LastWeekR)
			LRws = LRwb.active
			MR = LRws.max_row
			#print('last',MR)
			if MR==0: MR+=1
			for l in range(1,MR+1):
				#print(l,LRws.cell(row=l, column=1).value)
				if LRws.cell(row = l, column = C.Date_COL).value == ws.title:
					P = LRws.cell(row = l, column = C.Openprice_COL).value

					Rws.cell(row = C_maxrow, column = 2+count).value = score - P
					Rws.cell(row = C.NAME_ROW, column = 2+count).value = LastWeek
					break
			count += 1


	

#def insertRow(Iws,r,ws,A,Ticker_COL=0,Lista_COL=0,Kurs_COL=0,Namn_COL=0,Datum=0):
def insertRow(Into,IntoRow,From,FromRow,Kurs_COL=1,Ticker_COL=1,Lista_COL=1,Namn_COL=1,Datum=0):
	""" Insert a row """
	#move all rows below insertion down one.
	for a in range(Into.max_row,IntoRow-1,-1):
		for b in range(1,Into.max_column+1):
			Into.cell(row = a+1, column = b).value = Into.cell(row = a, column = b).value

	#insert every col
	for b in range(1,Into.max_column+1):
		if b == C.Date_COL:
			Into.cell(row = a, column = b).value = Datum
		elif b == C.Closeprice_COL:
			Into.cell(row = a, column = b).value = From.cell(row = FromRow, column = Kurs_COL).value
		else:
			Into.cell(row = a, column = b).value = ""

def importOneFromDaily(N, Nd):

	wb = openpyxl.load_workbook(Nd)
	ws = wb.active

	Iwb = openpyxl.load_workbook(N)
	Iws = Iwb.active

	Datum = Nd.split('_')[1]
	Datum = Datum.split('.')[0]
	Datum = datetime.datetime.strptime(Datum, '%Y-%M-%d')


	for allColl in range(1,ws.max_column+1):
		if ws.cell(row = 2, column = allColl).value == "Ticker":
			Ticker_COL = allColl
		if ws.cell(row = 2, column = allColl).value == "Lista":
			Lista_COL  = allColl
		if ws.cell(row = 1, column = allColl).value == "Aktiekurs":
			Kurs_COL   = allColl
		if ws.cell(row = 1, column = allColl).value == "Bolagsnamn":
			Namn_COL   = allColl
	for A in range(3, ws.max_row+1):
			temp_A = ws.cell(row = A, column = Ticker_COL).value

			first = temp_A.split(" ")[0]
			stript = temp_A.replace(" ","")

			#Om kortnamnet finns med i listan över ändrade aktier, så byt namn till det nya.
			for new in Changed:
				if first == new[1]:
					first = new[0]
				if stript == new[1]:
					stript = new[0]
			ReportName_F = str(first)+"-Price.xlsx"
			ReportName_S = str(stript)+"-Price.xlsx"
			found = False
			fileToImportTo = ""
			if os.path.isfile(C.DataDir+"/"+ReportName_F):
				#print("First: ", ReportName_F)
				found = True
				fileToImportTo = ReportName_F
				allFromDaily.append(ReportName_F)
			if os.path.isfile(C.DataDir+"/"+ReportName_S) and not found:
				#print("Stript: ", ReportName_S)
				fileToImportTo = ReportName_S
				found = True
				allFromDaily.append(ReportName_S)

			if found:
				for _row in range(2,Iws.max_row+1):
					#print(fileToImportTo+' '+str(Iws.cell(row = _row, column = C.Date_COL).value)+' '+str(Datum))
					if ((Iws.cell(row = _row, column = C.Date_COL).value - Datum).days) < 0:
						insertRow(Iws,_row,ws,A,Kurs_COL,Ticker_COL,Lista_COL,Namn_COL,Datum)
						break
def importFromDaily():
	allFromDaily = []
	Changed = []
	with open("ChagedNames.txt",'r') as f:
		a = f.readlines()
		for aa in a:
			aa = aa.rstrip()
			aa = aa.split(" ")
			Changed.append(aa)
			#ReportName_ = str(first)+"-Price.xlsx"
	for AllNames in glob.glob(C.DailyDir+"/Bors*xlsx"):
		#Namn = WB.split('/')[1]
		#Namn = Namn.split('-')[0]

		#print("MakeLastReports",Namn)

		wb = openpyxl.load_workbook(AllNames)
		ws = wb.active

		Datum = AllNames.split('_')[1]
		Datum = Datum.split('.')[0]
		Datum = datetime.datetime.strptime(Datum, '%Y-%M-%d')
		
		#find what column contains what
		for allColl in range(1,ws.max_column+1):
			if ws.cell(row = 2, column = allColl).value == "Ticker":
				Ticker_COL = allColl
			if ws.cell(row = 2, column = allColl).value == "Lista":
				Lista_COL  = allColl
			if ws.cell(row = 1, column = allColl).value == "Aktiekurs":
				Kurs_COL   = allColl
			if ws.cell(row = 1, column = allColl).value == "Bolagsnamn":
				Namn_COL   = allColl

		#För varje aktie
		for A in range(3, ws.max_row+1):
			temp_A = ws.cell(row = A, column = Ticker_COL).value

			first = temp_A.split(" ")[0]
			stript = temp_A.replace(" ","")

			#Om kortnamnet finns med i listan över ändrade aktier, så byt namn till det nya.
			for new in Changed:
				if first == new[1]:
					first = new[0]
				if stript == new[1]:
					stript = new[0]
			ReportName_F = str(first)+"-Price.xlsx"
			ReportName_S = str(stript)+"-Price.xlsx"
			found = False
			fileToImportTo = ""
			if os.path.isfile(C.DataDir+"/"+ReportName_F):
				#print("First: ", ReportName_F)
				found = True
				fileToImportTo = ReportName_F
				allFromDaily.append(ReportName_F)
			if os.path.isfile(C.DataDir+"/"+ReportName_S) and not found:
				#print("Stript: ", ReportName_S)
				fileToImportTo = ReportName_S
				found = True
				allFromDaily.append(ReportName_S)

			if found:
				#D = ws.cell(row = i, column = C.Date_COL).value
				Iwb = openpyxl.load_workbook(C.DataDir+"/"+fileToImportTo)
				Iws = Iwb.active
				print(C.DataDir+"/"+fileToImportTo)
				for _row in range(2,Iws.max_row+1):
					#print(fileToImportTo+' '+str(Iws.cell(row = _row, column = C.Date_COL).value)+' '+str(Datum))
					if ((Iws.cell(row = _row, column = C.Date_COL).value - Datum).days) < 0:
						insertRow(Iws,_row,ws,A,Ticker_COL,Lista_COL,Kurs_COL,Namn_COL,Datum)
						break
					#if (Iws.cell(row = 2, column = C.Date_COL).value - Datum)
				#S0tuff needs to go here
				#print(C.DataDir+"/"+fileToImportTo)
				#print(Iws.cell(row = 2, column = C.Date_COL).value)
				#InsertDatum = Iws.cell(row = 2, column = C.Date_COL).value
				#LastDateDiff = InsertDatum - Datum
				#if LastDateDiff < datetime.timedelta(hours = 12):
				#	insertRow(Iws,2,ws,A,Ticker_COL,Lista_COL,Kurs_COL,Namn_COL,Datum)
				#	#insertDailyData(Iws)
				#for koll in range(30,2,-1):
				#	today = Iws.cell(row = koll, column = C.Date_COL).value
				#	yesterday = Iws.cell(row = koll+1, column = C.Date_COL).value
				#	if ( today - Datum ) > datetime.timedelta(hours = 12) and ( yesterday - Datum ) < datetime.timedelta(hours = 12):
				#		insertRow(Iws,koll+1,ws,A,Ticker_COL,Lista_COL,Kurs_COL,Namn_COL,Datum) 
				Iwb.save(C.DataDir+"/"+fileToImportTo)
				#print(Datum - datetime.datetime.today())

			#print("First: ", ReportName_F)
	"""
	for AllNames in glob.glob(DataDir+"/*xlsx"):
		AllNames=AllNames.split("/")[1]
		if AllNames not in allFromDaily:
			print(AllNames)
	"""
	#print("AFD",len(allFromDaily))
	#print(allFromDaily)
