import openpyxl
#from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import math
import glob
import datetime
import time
import os

import Swingtrader_constants as C

def copyFromDownloads():
	""" Copy from Downloads to Inbox """
	for ORG in glob.glob("/home/sid/Downloads/*-Price.xls"):
		try: os.makedirs(C.DataDir+'/Inbox/')
		except: pass
		TO = ORG.split('/')[-1]
		TO = C.DataDir+'/'+TO
		os.system('mv '+ORG+' '+TO)
		#os.system('rm '+ORG)
	makeXLSX()

def makeXLSX():
	""" Make xlsx-files, move to DataDir, and remove original """
	os.system('libreoffice --convert-to xlsx '+C.DataDir+'/*.xls --headless')

	for ORG in glob.glob("*-Price.xlsx"):
		print(ORG)
		TO = ORG.split('/')[-1]
		TO = C.DataDir+'/'+TO
		os.system('mv '+ORG+' '+TO)
		#os.system('rm '+ORG)
	for ORG in glob.glob(C.DataDir+"/*-Price.xls"):
		os.system('rm '+ORG)

def makeReport(N):
	""" make report for a given stock """
	MAXlenTIME = 10000 # maximum amount of days back in time

	try: os.makedirs(C.DataDir+'/Reports/')
	except: pass

	Namn = N.split('/')[-1]
	Namn = Namn.split('-')[0]

	wb = openpyxl.load_workbook(N)
	ws = wb.active

	if ws.cell(row = C.NAME_ROW, column = C.MA15_COL).value != "MA15":
		return 'MA15'
	if ws.cell(row = C.NAME_ROW, column = C.DMA15_COL).value != "DMA15":
		return 'DMA15'
	
	startrow = min(MAXlenTIME,ws.max_row)
	# From the last row, to the first row:
	for i in range(startrow, 1, -1):
		D = ws.cell(row = i, column = C.Date_COL).value
		
		# Get the next day, if there is one
		# Else sett the same day
		try: ND = ws.cell(row = i-1, column = C.Date_COL).value
		except: ND = ws.cell(row = i, column = C.Date_COL).value

		# if this day has a greater number, it is the last of the week
		##if D.weekday() == C.ReportWeekday and ws.cell(row = i, column = C.MA200_COL).value != 'NaN':
		if D.weekday() >= ND.weekday() and ws.cell(row = i, column = C.MA200_COL).value != 'NaN':
			YearDay = str(D.year)+'_'+str(D.month)+'_'+str(D.day)

			ReportName = C.DataDir+'/Reports/'+str(YearDay)+'_report.xlsx'

			if not os.path.isfile(ReportName):
				Rwb = openpyxl.Workbook()
			else:
				Rwb = openpyxl.load_workbook(ReportName)
			Rws = Rwb.active

			C_maxrow = Rws.max_row
			Cval  = ws.cell(row = i, column = C.Closeprice_COL).value
			SHORT_TIMEFRAME  = ws.cell(row = i, column = C.MA15_COL).value
			MEDIUM_TIMEFRAME = ws.cell(row = i, column = C.MA50_COL).value
			LONG_TIMEFEFRAME = ws.cell(row = i, column = C.MA200_COL).value

			score = 0

			if SHORT_TIMEFRAME > MEDIUM_TIMEFRAME and Cval > MEDIUM_TIMEFRAME:
				score += 1
			if MEDIUM_TIMEFRAME > LONG_TIMEFEFRAME and Cval > LONG_TIMEFEFRAME:
				score += 2
			Rws.cell(row = C_maxrow+1, column = C.Date_COL).value = ws.title
			Rws.cell(row = C_maxrow+1, column = C.Openprice_COL).value = score
			Rws.cell(row = C.NAME_ROW, column = C.Openprice_COL).value = YearDay
			Rws.cell(row = C.NAME_ROW, column = C.Date_COL).value = "Name"
			
			try: importWeekHistory(ws, Rws, D, score)
			except: print("Could not import history for ", ReportName)
			Rwb.save(ReportName)

def importWeekHistory(ws, Rws, D, score):
	""" Find the report for last week and insert to Rws """
	count = 1
	C_maxrow = Rws.max_row
	for jj in [2,9]:
		LW = D - datetime.timedelta(days = jj)
		LastWeek = str(LW.year)+'_'+str(LW.month)+'_'+str(LW.day)
		LastWeekR = C.DataDir+'/Reports/'+str(LastWeek)+'_report.xlsx'
		if os.path.isfile(LastWeekR):
			
			LRwb = openpyxl.load_workbook(LastWeekR)
			LRws = LRwb.active
			
			Rws.cell(row = C.NAME_ROW, column = 2+count).value = LastWeek
			
			MR = LRws.max_row
			if MR==0: MR+=1
			
			for l in range(1,MR+1):
				#print(l,LRws.cell(row=l, column=1).value)
				if LRws.cell(row = l, column = C.Date_COL).value == ws.title:
					P = LRws.cell(row = l, column = C.Openprice_COL).value

					Rws.cell(row = C_maxrow, column = 2+count).value = score - P
					break
			count += 1