# def makeAllReports(WB):
# def makeLastReport(WB):

import openpyxl
#from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import math
import glob
import datetime
import time
import os

import Swingtrader_constants as C

from TA_functions     import *
from Import_functions import *
from Report_functions import *

# /home/sid/Documents


def makeReport(N):
	#global max_maxrow, DataDir

	try: os.makedirs(C.DataDir+'/Reports/')
	except: pass

	Namn = N.split('/')[-1]
	Namn = Namn.split('-')[0]

	#print("MakeAllReports",Namn)
	wb = openpyxl.load_workbook(N)
	#wb = openpyxl.load_workbook(WB)
	ws = wb.active

	if ws.cell(row = C.NAME_ROW, column = C.MA15_COL).value != "MA15":
		print(ws.cell(row = C.NAME_ROW, column = C.MA15_COL).value)
		return 'MA15'
	if ws.cell(row = C.NAME_ROW, column = C.DMA15_COL).value != "DMA15":
		return 'DMA15'
	
	for i in range(ws.max_row, 1, -1):
		D = ws.cell(row = i, column = C.Date_COL).value
		if D.weekday() == C.ReportWeekday and ws.cell(row = i, column = C.MA200_COL).value != 'NaN':
			YearDay = str(D.year)+'_'+str(D.month)+'_'+str(D.day)

			ReportName = C.DataDir+'/Reports/'+str(YearDay)+'_report.xlsx'

			if not os.path.isfile(ReportName):
				Rwb = openpyxl.Workbook()
			else:
				Rwb = openpyxl.load_workbook(ReportName)
			Rws = Rwb.active

			C_maxrow = Rws.max_row
			Cval  = ws.cell(row = i, column = C.Closeprice_COL).value
			SHORT_TIMEFRAME  = ws.cell(row = i, column = C.MA15_COL).value
			MEDIUM_TIMEFRAME = ws.cell(row = i, column = C.MA50_COL).value
			LONG_TIMEFEFRAME = ws.cell(row = i, column = C.MA200_COL).value

			score = 0

			if SHORT_TIMEFRAME > MEDIUM_TIMEFRAME and Cval > MEDIUM_TIMEFRAME:
				score += 1
			if MEDIUM_TIMEFRAME > LONG_TIMEFEFRAME and Cval > LONG_TIMEFEFRAME:
				score += 2
			Rws.cell(row = C_maxrow+1, column = C.Date_COL).value = ws.title
			Rws.cell(row = C_maxrow+1, column = C.Openprice_COL).value = score
			Rws.cell(row = C.NAME_ROW, column = C.Openprice_COL).value = YearDay
			Rws.cell(row = C.NAME_ROW, column = C.Date_COL).value = "Name"
			
			try: importWeekHistory(ws, Rws, D, score)
			except: print("Could not import history for ", ReportName)
			Rwb.save(ReportName)
def makeAllReports(N):
	#global max_maxrow, DataDir

	try: os.makedirs(C.DataDir+'/Reports/')
	except: pass

	Namn = N.split('/')[1]
	Namn = Namn.split('-')[0]

	#print("MakeAllReports",Namn)
	wb = openpyxl.load_workbook(N)
	#wb = openpyxl.load_workbook(WB)
	ws = wb.active

	if ws.cell(row = C.NAME_ROW, column = C.MA15_COL).value != "MA15":
		print("SHOULD NOT BE HERE!!!!")
		allaGrafer.printMAtoCurrentWB(wb,Namn,C.DataDir)
		allaGrafer.DeriveAndPrintDMAtoCurrentWB(wb,Namn,C.DataDir)

		if ws.max_row > C.max_maxrow:
			C.max_maxrow = ws.max_row

	for i in range(ws.max_row, 1, -1):
		D = ws.cell(row = i, column = C.Date_COL).value
		if D.weekday() == C.ReportWeekday and ws.cell(row = i, column = C.MA200_COL).value != 'NaN':
			YearDay = str(D.year)+'_'+str(D.month)+'_'+str(D.day)

			ReportName = C.DataDir+'/Reports/'+str(YearDay)+'_report.xlsx'

			if not os.path.isfile(ReportName):
				Rwb = openpyxl.Workbook()
			else:
				Rwb = openpyxl.load_workbook(ReportName)
			Rws = Rwb.active

			C_maxrow = Rws.max_row
			Cval  = ws.cell(row = i, column = C.Closeprice_COL).value
			SHORT_TIMEFRAME  = ws.cell(row = i, column = C.MA15_COL).value
			MEDIUM_TIMEFRAME = ws.cell(row = i, column = C.MA50_COL).value
			LONG_TIMEFEFRAME = ws.cell(row = i, column = C.MA200_COL).value

			score = 0

			if SHORT_TIMEFRAME > MEDIUM_TIMEFRAME and Cval > MEDIUM_TIMEFRAME:
				score += 1
			if MEDIUM_TIMEFRAME > LONG_TIMEFEFRAME and Cval > LONG_TIMEFEFRAME:
				score += 2
			Rws.cell(row = C_maxrow+1, column = C.Date_COL).value = ws.title
			Rws.cell(row = C_maxrow+1, column = C.Openprice_COL).value = score
			Rws.cell(row = C.NAME_ROW, column = C.Openprice_COL).value = YearDay
			Rws.cell(row = C.NAME_ROW, column = C.Date_COL).value = "Name"
			
			importHistory(ws, Rws, D, score)
			Rwb.save(ReportName)
			
def makeLastReport(WB):
	#global max_maxrow, DataDir

	try: os.makedirs(C.DataDir+'/Reports/')
	except: pass

	Namn = WB.split('/')[1]
	Namn = Namn.split('-')[0]

	#print("MakeLastReports",Namn)

	wb = openpyxl.load_workbook(WB)
	ws = wb.active

	if ws.cell(row = C.NAME_ROW, column = C.MA15_COL).value != "MA15":
		allaGrafer.printMAtoCurrentWB(wb,Namn,C.DataDir)
		allaGrafer.DeriveAndPrintDMAtoCurrentWB(wb,Namn,C.DataDir)

		if ws.max_row > C.max_maxrow:
			C.max_maxrow = ws.max_row

	#for i in range(ws.max_row, 1, -1):
	i = 2
	D = ws.cell(row = i, column = C.Date_COL).value
	if ws.cell(row = i, column = C.MA200_COL).value != 'NaN':
		YearDay = str(D.year)+'_'+str(D.month)+'_'+str(D.day)

		ReportName = C.DataDir+'/Reports/'+str(YearDay)+'_report.xlsx'

		if not os.path.isfile(ReportName):
			Rwb = openpyxl.Workbook()
		else:
			Rwb = openpyxl.load_workbook(ReportName)
		Rws = Rwb.active

		C_maxrow = Rws.max_row
		Cval  = ws.cell(row = i, column = C.Closeprice_COL).value
		SHORT_TIMEFRAME  = ws.cell(row = i, column = C.MA15_COL).value
		MEDIUM_TIMEFRAME = ws.cell(row = i, column = C.MA50_COL).value
		LONG_TIMEFEFRAME = ws.cell(row = i, column = C.MA200_COL).value

		score = 0

		if SHORT_TIMEFRAME > MEDIUM_TIMEFRAME and Cval > MEDIUM_TIMEFRAME:
			score += 1
		if MEDIUM_TIMEFRAME > LONG_TIMEFEFRAME and Cval > LONG_TIMEFEFRAME:
			score += 2
		Rws.cell(row = C_maxrow+1, column = C.Date_COL).value = ws.title
		Rws.cell(row = C_maxrow+1, column = C.Openprice_COL).value = score
		Rws.cell(row = C.NAME_ROW, column = C.Openprice_COL).value = YearDay
		Rws.cell(row = C.NAME_ROW, column = C.Date_COL).value = "Name"
		
		count = 1
		C_maxrow = Rws.max_row
		for jj in range(1,30):
			LW = D - datetime.timedelta(days = jj)
			LastWeek = str(LW.year)+'_'+str(LW.month)+'_'+str(LW.day)
			LastWeekR = C.DataDir+'/Reports/'+str(LastWeek)+'_report.xlsx'
			if os.path.isfile(LastWeekR):
				
				LRwb = openpyxl.load_workbook(LastWeekR)
				LRws = LRwb.active
				MR = LRws.max_row
				#print('last',MR)
				if MR==0: MR+=1
				for l in range(1,MR+1):
					#print(l,LRws.cell(row=l, column=1).value)
					if LRws.cell(row = l, column = C.Date_COL).value == ws.title:
						P = LRws.cell(row = l, column = C.Openprice_COL).value

						Rws.cell(row = C_maxrow,   column = 2+count).value = score - P
						Rws.cell(row = C.NAME_ROW, column = 2+count).value = LastWeek
						break
				if jj > 3:
					break	
				count += 1
		Rwb.save(ReportName)

def copyFromDownloads():
	for ORG in glob.glob("/home/sid/Downloads/*-Price.xls"):
		#os.system('cls||clear')
		try: os.makedirs(C.DataDir+'/Inbox/')
		except: pass
		print("one")
		TO = ORG.split('/')[-1]
		#TO = C.DataDir+'/Inbox/'+TO
		TO = C.DataDir+'/'+TO
		#print("one")
		os.system('cp '+ORG+' '+TO)
		os.system('rm '+ORG)
		#print('cp '+ORG+' '+TO)
		#print("one")
		#print(al)
	makeXLSX()
def makeXLSX():
	print("two")
	os.system('libreoffice --convert-to xlsx '+C.DataDir+'/*.xls --headless')
	#os.system('libreoffice --convert-to xlsx '+C.DataDir+'/Inbox/*.xls --headless')
	#print(glob.glob("*-Price.xlsx"))
	
	for ORG in glob.glob("*-Price.xlsx"):
		print(ORG)
		TO = ORG.split('/')[-1]
		#TO = C.DataDir+'/Inbox/'+TO
		TO = C.DataDir+'/'+TO
		os.system('cp '+ORG+' '+TO)
		os.system('rm '+ORG)
	#for ORG in glob.glob(C.DataDir+"/Inbox/*-Price.xls"):
	for ORG in glob.glob(C.DataDir+"/*-Price.xls"):
		os.system('rm '+ORG)
	#copyFromInbox()
	
def copyFromInbox():
	for ORG in glob.glob(C.DataDir+"/Inbox/*-Price.xlsx"):
		TO = ORG.split('/')[-1]
		TO =  TO.split('-')[0]

		wb = openpyxl.load_workbook(ORG)
		ws = wb.active

		MasterDir = C.DataDir+'/'+TO+'-Master.xlsx'

		if not os.path.isfile(MasterDir):
			Mwb = openpyxl.Workbook()
		else:
			Mwb = openpyxl.load_workbook(MasterDir)
		Mws = Rwb.active

		for R in range(1,ws.max_row):
			pass#ws.

def EqualRow(ws1,ws2,r1,r2):
	if ws1.cell(row = r1, column = 1).value != ws2.cell(row = r1, column = 1).value: return False
	if ws1.cell(row = r1, column = 1).value != ws2.cell(row = r1, column = 2).value: return False
	if ws1.cell(row = r1, column = 1).value != ws2.cell(row = r1, column = 3).value: return False
	if ws1.cell(row = r1, column = 1).value != ws2.cell(row = r1, column = 4).value: return False
	if ws1.cell(row = r1, column = 1).value != ws2.cell(row = r1, column = 5).value: return False
	if ws1.cell(row = r1, column = 1).value != ws2.cell(row = r1, column = 6).value: return False
	return True