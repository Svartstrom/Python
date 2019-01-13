import unittest
import os
import openpyxl

from TA_functions     import *
from Import_functions import *
from Report_functions import *

class TestInsertRow(unittest.TestCase):
	def setUp(self):
		os.system('cp Testdata/org_Borsdata_2018-01-02.xlsx Testdata/Borsdata_2018-01-02.xlsx')
		os.system('cp Testdata/org_AAK-Price.xlsx Testdata/AAK-Price.xlsx')
	def tearDown(self):
		os.system('rm Testdata/Borsdata_2018-01-02.xlsx')
		os.system('rm Testdata/AAK-Price.xlsx')
	def test_insert(self):
		wb = openpyxl.load_workbook('Testdata/AAK-Price.xlsx')
		ws = wb.active
		wbF = openpyxl.load_workbook('Testdata/Borsdata_2018-01-02.xlsx')
		wsF = wbF.active

		#insertRow(TO,2,FROM,14,Kurs_COL,Ticker_COL,Lista_COL,Namn_COL,Datum)
		insertRow(ws,2,wsF,14,14,1,1,1,2017)          
		wb.save('Testdata/AAK-Price.xlsx')      
		value = ws.cell(row = 2, column = C.Closeprice_COL).value
		self.assertEqual(value,694.0)
		#insertRow(Iws,_row,ws,A)

class TestImportHistory(unittest.TestCase):
	pass
def tester():
	unittest.main()

if __name__ == '__main__':
	unittest.main()
