# def printFullMAtoCurrentWB(wb,DataDir):
# def printOneMAtoCurrentWB(wb,r,DataDir):
# def DeriveAndPrintFullDMAtoCurrentWB(wb,DataDir):
# def PlotAgainstIndex(ws,Namn,Index):

import openpyxl
#from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import math
import glob
import datetime
import time
import os

import Swingtrader_constants as C

def deleteFullMAandDMA(N):
	wb = openpyxl.load_workbook(N)
	ws = wb.active

	for R in range(1,ws.max_row+1):
		ws.cell(row = R, column = C.MA15_COL  ).value   = ""
		ws.cell(row = R, column = C.MA50_COL  ).value   = ""
		ws.cell(row = R, column = C.MA200_COL ).value  = ""
		ws.cell(row = R, column = C.DMA15_COL ).value  = ""
		ws.cell(row = R, column = C.DMA50_COL ).value  = ""
		ws.cell(row = R, column = C.DMA200_COL).value = ""
	wb.save(N)

def printFullMAtoCurrentWB(N,DataDir):
	# Sparar inte WB!
	#import Swingtrader_constants

	SHORT_TIMEFRAME  	= []
	MEDIUM_TIMEFRAME  = []
	LONG_TIMEFRAME 	= []
	wb = openpyxl.load_workbook(N)
	ws = wb.active
	ws.cell(row = C.NAME_ROW, column = C.MA15_COL ).value  = "MA"+str(C.ShortAmountDays)
	ws.cell(row = C.NAME_ROW, column = C.MA50_COL ).value  = "MA"+str(C.MediumAmountDays)
	ws.cell(row = C.NAME_ROW, column = C.MA200_COL).value = "MA" +str(C.LongAmountDays)

	for i in range(ws.max_row,1,-1):
		Cval = ws.cell(row = i, column = C.Closeprice_COL).value

		SHORT_TIMEFRAME.append(Cval)
		if len(SHORT_TIMEFRAME) >= C.ShortAmountDays:
			S = sum(SHORT_TIMEFRAME) / C.ShortAmountDays
			ws.cell(row = i, column = C.MA15_COL).value = S
			del(SHORT_TIMEFRAME[0])
		else:
			ws.cell(row = i, column = C.MA15_COL).value = 'NaN'

		MEDIUM_TIMEFRAME.append(Cval)#ws.cell(row=i,column=5).value)
		if len(MEDIUM_TIMEFRAME) > C.MediumAmountDays:
			S = sum(MEDIUM_TIMEFRAME) / C.MediumAmountDays
			ws.cell(row = i, column = C.MA50_COL).value = S
			del(MEDIUM_TIMEFRAME[0])
		else:
			ws.cell(row = i, column = C.MA50_COL).value = 'NaN'

		LONG_TIMEFRAME.append(Cval)
		if len(LONG_TIMEFRAME) > C.LongAmountDays:
			S = sum(LONG_TIMEFRAME) / C.LongAmountDays
			ws.cell(row = i, column = C.MA200_COL).value = S
			del(LONG_TIMEFRAME[0])
		else:
			ws.cell(row = i, column = C.MA200_COL).value = 'NaN'
	wb.save(N)

def printOneMAtoCurrentWB(wb,r,DataDir):
	# Sparar inte WB!
	#import Swingtrader_constants

	C.SHORT_TIMEFRAME  	= []
	C.MEDIUM_TIMEFRAME  = []
	C.LONG_TIMEFRAME 	= []
	ws = wb.active
	ws.cell(row = C.NAME_ROW, column = C.MA15_COL).value  = "MA"+str(C.ShortAmountDays)
	ws.cell(row = C.NAME_ROW, column = C.MA50_COL).value  = "MA"+str(C.MediumAmountDays)
	ws.cell(row = C.NAME_ROW, column = C.MA200_COL).value = "MA"+str(C.LongAmountDays)

	for i in range(r+200,r,-1):
		Cval = ws.cell(row = i, column = C.Closeprice_COL).value

		C.SHORT_TIMEFRAME.append(Cval)
		if len(C.SHORT_TIMEFRAME) >= C.ShortAmountDays:
			S = sum(C.SHORT_TIMEFRAME) / C.ShortAmountDays
			ws.cell(row = i, column = C.MA15_COL).value = S
			del(C.SHORT_TIMEFRAME[0])
		else:
			pass#ws.cell(row = i, column = C.MA15_COL).value = 'NaN'

		C.MEDIUM_TIMEFRAME.append(Cval)#ws.cell(row=i,column=5).value)
		if len(C.MEDIUM_TIMEFRAME) > C.MediumAmountDays:
			S = sum(C.MEDIUM_TIMEFRAME) / C.MediumAmountDays
			ws.cell(row = i, column = C.MA50_COL).value = S
			del(C.MEDIUM_TIMEFRAME[0])
		else:
			ws.cell(row = i, column = C.MA50_COL).value = 'NaN'

		C.LONG_TIMEEFRAME.append(Cval)
		if len(C.LONG_TIMEFEFRAME) > C.LongAmountDays:
			S = sum(C.LONG_TIMEFREFRAME) / C.LongAmountDays
			ws.cell(row = i, column = C.MA200_COL).value = S
			del(C.LONG_TIMEFREFRAME[0])
		else:
			ws.cell(row = i, column = C.MA200_COL).value = 'NaN'
	#wb.save(DataDir+"/"+Namn+"-Price.xlsx")

def DeriveAndPrintFullDMAtoCurrentWB(N,DataDir):
	wb = openpyxl.load_workbook(N)
	ws = wb.active
	if ws.cell(row = C.NAME_ROW, column = C.MA15_COL).value != "MA15": return -1

	D_SHORT_TIMEFRAME  = []
	dma50  = []
	dma200 = []	

	ws.cell(row = C.NAME_ROW, column = C.DMA15_COL).value = "DMA15"
	ws.cell(row = C.NAME_ROW, column = C.DMA50_COL).value = "DMA50"
	ws.cell(row = C.NAME_ROW, column = C.DMA200_COL).value = "DMA200"

	prevMA15  = ws.cell(row = ws.max_row, column = C.MA15_COL ).value
	prevMA50  = ws.cell(row = ws.max_row, column = C.MA50_COL ).value
	prevMA200 = ws.cell(row = ws.max_row, column = C.MA200_COL).value

	for i in range(ws.max_row,1,-1):
		if ws.cell(row = i, column = C.MA15_COL).value != 'NaN':
			if prevMA15 == 'NaN':
				prevMA15 = ws.cell(row = i, column = C.MA15_COL).value
				ws.cell(row = i, column = C.DMA15_COL).value = 'NaN'
			else:
				Cval = ws.cell(row = i, column = C.MA15_COL).value
				ws.cell(row = i, column = C.DMA15_COL).value = Cval - prevMA15
				prevMA15 = Cval
		else:	
			ws.cell(row = i, column = C.DMA15_COL).value = 'NaN'

		if ws.cell(row = i, column = C.MA50_COL).value != 'NaN':
			if prevMA50 == 'NaN':
				prevMA50 = ws.cell(row = i, column = C.MA50_COL).value
				ws.cell(row = i, column = C.DMA50_COL).value = 'NaN'
			else:
				Cval = ws.cell(row = i, column = C.MA50_COL).value
				ws.cell(row = i, column = C.DMA50_COL).value = Cval - prevMA50
				prevMA50 = Cval
		else:	
			ws.cell(row = i, column = C.DMA50_COL).value = 'NaN'

		if ws.cell(row = i, column = C.MA200_COL).value != 'NaN':
			if prevMA200 == 'NaN':
				prevMA200 = ws.cell(row = i, column = C.MA200_COL).value
				ws.cell(row = i, column = C.DMA200_COL).value = 'NaN'
			else:
				Cval = ws.cell(row = i, column = C.MA200_COL).value
				ws.cell(row = i, column = C.DMA200_COL).value = Cval - prevMA200
				prevMA200 = Cval
		else:	
			ws.cell(row = i, column = C.DMA200_COL).value = 'NaN'
	wb.save(N)

def PlotAgainstIndex(ws,Namn,Index):
	global avg_earning, avg_alpha, wbMaster, wsMaster
	#wbMaster = openpyxl.Workbook("Data/Master.xlsx")
	#wsMaster = wbMaster.active

	C = [1]
	temp = []
	#print(type(temp))
	temp.append(1)
	#print(type(temp))
	ma15=[]
	ma50=[]
	ma200=[]
	MA15=[0 for a in range(15)] 
	MA50=[0 for a in range(50)]
	MA200=[0 for a in range(200)]
	saldo=[1]
	saldo200 =[1]
	avg = [1]
	for i in range(ws.max_row,1,-1):

		D = ws.cell(row=i, column = 1).value

		wsMaster.cell(row=i, column = 1).value = D

		#THIS NEEDS TO BE REMADE!!!!
		# Den ska plotta ma15, ma50, ma200 i varje graf
		# 
		#print(type(temp))
		C.append(    ws.cell(row=i,column=5).value)
		ma15.append( ws.cell(row=i,column=5).value)
		ma50.append( ws.cell(row=i,column=5).value)
		ma200.append(ws.cell(row=i,column=5).value)
		if len(ma15) > 14:
			S = sum(ma15) / 15
			#ws.cell(row=i,column=7).value = S
			MA15.append(S)
			del(ma15[0])
		else:
			MA15.append(float('NaN'))
		if len(ma50) > 49:
			S = sum(ma50) / 50
			#ws.cell(row=i,column=8).value = S
			MA50.append(S)
			del(ma50[0])
		else:
			MA50.append(float('NaN'))
		if len(ma200) > 199:
			S = sum(ma200) / 200
			#ws.cell(row=i,column=9).value = S
			MA200.append(S)
			del(ma200[0])
		else:
			MA200.append(float('NaN'))
		temp = saldo[-1]
		temp2 = avg[-1]
		temp200=saldo200[-1]
		#print("C",C)
		
		try:
			wsMaster.cell(row=i, column = 2).value += 1
		except TypeError:
			wsMaster.cell(row=i, column = 2).value = 1

		if MA15[-1] > MA50[-1] and C[-1] > MA200[-1]:# and MA50[-1] > MA200[-1]:
			#temp -= 500
			try:
				wsMaster.cell(row=i, column = 3).value += 1
			except TypeError:
				wsMaster.cell(row=i, column = 3).value = 1
			temp  += (saldo[-1]/2) * ((C[-1]/C[-2])-1)
			#print("C",C[-1],C[-2],C)
		if MA50[-1] > MA200[-1] and C[-1] > MA200[-1]:
			#temp -= 500
			try:
				wsMaster.cell(row=i, column = 4).value += 1
			except TypeError:
				wsMaster.cell(row=i, column = 4).value = 1
			temp  += (saldo[-1]/2) * ((C[-1]/C[-2])-1)
			temp200  += (saldo[-1]) * ((C[-1]/C[-2])-1)
		temp2  += (avg[-1]) * ((C[-1]/C[-2])-1)
		#avg.append(temp2)
		saldo.append(temp)
		saldo200.append(temp200)

	with open(RepotdDoc,'a') as S:
		S.write(ws.title+'\n')
		if MA15[-1] > MA50[-1] and C[-1] > MA200[-1]:# and MA50[-1] > MA200[-1]:
			S.write('Köp kort, Ma50 = %s\n'%str(MA50[-1]))
		else:
			S.write('Sälj kort, Ma50 = %s\n'%str(MA50[-1]))
		if MA50[-1] > MA200[-1] and C[-1] > MA200[-1]:
			#temp -= 500
			S.write('Köp lång, Ma200 = %s\n\n'%str(MA200[-1]))
		else:
			S.write('Sälj lång, Ma200 = %s\n\n'%str(MA200[-1]))
	X = [ii for ii in range(len(C))]
	Index = Index[-len(C):]
	#plt.figure()
	#plt.subplot()
	
	Tindex = [(C[0]/Index[0])*ii for ii in Index]
	Index = Tindex

	P = plt.figure()
	plt.subplot(111)
	#print(len(X))
	#print(len(C))
	#print(len(Index))
	#print(len(saldo200))

	dD = D-datetime.datetime.today()#.strptime(D,"%Y-%m-%d")
	dD = datetime.datetime(2017,1,1)-D
	dD = dD.days
	try:
		temp200 = [C[dD]/saldo200[dD] * ii for ii in saldo200]
		tempIndex = [C[dD]/Index[dD] * ii for ii in Index] 
		saldo200 = temp200
		Index = tempIndex
	except:
		pass
	try:
		plt.plot(X,C,X,Index,X,saldo200,X[dD],C[dD],'*')
	except:
		plt.plot(X,C,X,saldo200)
	P.savefig("Data/%s-Figure.png"%Namn)#,dpi=72)
	plt.close(P)
	try:
		print("Överavkastning: ",saldo200[-1]/C[-1])
		print("Avkastning: ", saldo200[-1]/saldo200[dD])
		avg_earning.append(saldo200[-1]/saldo200[dD])
		avg_alpha.append(saldo200[-1]/C[-1])
	except:
		pass
	print("")
	wbMaster.save("Data/Master.xlsx")
	#print("")
	#Print("Namn: ",Namn)