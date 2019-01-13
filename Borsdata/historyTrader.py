import Swingtrader_constants as C
import datetime
import os
import openpyxl
import glob
import math
import matplotlib.pyplot as plt

global money, stocks, length
money  = [100000,0]
stocks = []
length = [0,0]

# Nedan följer uppbyggnaden av "stocks"
## Namn  - Namn på aktien
## Längd - längd på innehavet, 1 - kort, 2 - lång
## Antal - Antal aktier
## Pris  - Pris för en aktie
# [Namn,längd,antal,pris]

def historyTrader():
	global money, stocks
	D = datetime.datetime.today()

	Iwb = openpyxl.load_workbook(C.DataDir+"/AAA_index.xlsx")
	Iws = Iwb.active

	if not os.path.isfile(C.DataDir+'/Reports/monneyReport.xlsx'):
		Mwb = openpyxl.Workbook()
	else:
		Mwb = openpyxl.load_workbook(C.DataDir+'/Reports/monneyReport.xlsx')
	Mws = Mwb.active

	Mws.cell(row = 1, column = 1).value = money[0]
	Mws.cell(row = 1, column = 2).value = money[1]
	# how many days back in time to start looking at reports
	for jj in range(50,0,-1):#range(5407,0,-1):
		#reset the amount in stocks for each day
		money[1] = 0

		#go back jj amount of days from today, create a string from that date, and fid the coresponding file
		LW = D - datetime.timedelta(days = jj)
		LastWeek = str(LW.year)+'_'+str(LW.month)+'_'+str(LW.day)
		LastWeekR = C.DataDir+'/Reports/'+str(LastWeek)+'_report.xlsx'

		#does that file exist?
		if os.path.isfile(LastWeekR):
			print(LastWeekR)
			# open the file
			wb = openpyxl.load_workbook(LastWeekR)
			ws = wb.active
			#sell all stocks recomended in the file
			sellStocks(ws,Iws,jj,LW)
			#buy all stocks recomended in the file
			buyStocks(ws,Iws,jj,LW)

		#go through all stocks and save the value
		for index,S in enumerate(stocks):
			money[1] += S[2]*S[3]

		print(str(math.floor(money[0])),str(math.floor(money[1])),str(math.floor(money[0]+money[1])))
		temp = Mws.max_row + 1
		
		#Save the amount of liquids and stocks 
		Mws.cell(row = temp, column = 1).value = money[0]
		Mws.cell(row = temp, column = 2).value = money[1]

	# We are done
	# Sell all stocks
	sellALLStocks(ws,Iws,jj,LW)
	
	if len(stocks) > 0:
		print("Not all stocks where sold")
		for a in stocks: print(a)

	#print the money 						stocks 						and total
	print( str(math.floor(money[0])) , str(math.floor(money[1])) , str(math.floor(money[0]+money[1])) )
	
	#one last save
	temp = Mws.max_row + 1
	Mws.cell(row = temp, column = 1).value = money[0]
	Mws.cell(row = temp, column = 2).value = money[1]

	# Acctual save of the excell
	Mwb.save(C.DataDir+'/Reports/monneyReport.xlsx')

def sellStocks(ws,Iws,jj,LW):
	global money, stocks,length
	if len(stocks) == 0:
		return
	for R in range(2,ws.max_row):
		

		if ws.cell(row = R, column = 3).value in [-1, -2, -3]:
			for index,S in enumerate(stocks):
				if S[0] == ws.cell(row = R, column = 1).value:
					if ws.cell(row = R, column = 3).value != -S[1]:
						#print("CONTINUE: ",ws.cell(row = R, column = 3).value, S[1])
						continue
					for I in range(1,Iws.max_row):
						iii=I
						if Iws.cell(row = I, column = 2).value == ws.cell(row = R, column = 1).value:
							break
					Dwb = openpyxl.load_workbook(C.DataDir+"/"+Iws.cell(row = iii, column = 1).value)
					Dws = Dwb.active
					checkDate = min(jj,ws.max_row)
					Date = Dws.cell(row = checkDate, column = 1).value
					if Date == 'Date': return
					print("Date LW",Date,LW)
					print(C.DataDir+"/"+Iws.cell(row = iii, column = 1).value)
					while ((LW -Date).days >= 0):
						checkDate -= 1
						Date = Dws.cell(row = checkDate, column = 1).value
					checkDate -= 1
					if checkDate == 1: return
					MondayValue = Dws.cell(row = checkDate, column = 2).value

					money[1] -= MondayValue * S[2]
					money[0] += MondayValue * S[2] * 0.9975

					length[0] += (S[4] - Dws.cell(row = checkDate, column = 1).value).days
					length[1] += 1
					#print("Sell "+str(S[0])+"with "+str((MondayValue * S[2] * 0.9925/(S[3] * S[2] * 0.9925)))+" % profit")
					del(stocks[index])
def sellALLStocks(ws,Iws,jj,LW):
	global money, stocks,length
	if len(stocks) == 0:
		return
	while len(stocks)>0:
		for R in range(2,ws.max_row):
			if True:#ws.cell(row = R, column = 3).value in [-1, -2, -3]:
				for index,S in enumerate(stocks):
					#print(S[0],ws.cell(row = R, column = 1).value)
					if S[0] == ws.cell(row = R, column = 1).value:
						if False:#ws.cell(row = R, column = 3).value != -S[1]:
							#print("CONTINUE: ",ws.cell(row = R, column = 3).value, S[1])
							continue
						for I in range(1,Iws.max_row):
							iii=I
							if Iws.cell(row = I, column = 2).value == ws.cell(row = R, column = 1).value:
								break
						Dwb = openpyxl.load_workbook(C.DataDir+"/"+Iws.cell(row = iii, column = 1).value)
						Dws = Dwb.active

						MondayValue = Dws.cell(row = 2, column = 2).value

						money[1] -= MondayValue * S[2]
						money[0] += MondayValue * S[2] * 0.9975

						length[0] += (S[4] - Dws.cell(row = 2, column = 1).value).days
						length[1] += 1
						#print("Sell "+str(S[0])+"with "+str((MondayValue * S[2] * 0.9925/(S[3] * S[2] * 0.9925)))+" % profit")
						del(stocks[index])
def buyStocks(ws,Iws,jj,LW):
	global money, stocks,length
	for R in range(2,ws.max_row):
		if ws.cell(row = R, column = 3).value in [1, 2, 3]:

			for I in range(1,Iws.max_row):
				iii=I
				if Iws.cell(row = I, column = 2).value == ws.cell(row = R, column = 1).value:
					break
			try:Dwb = openpyxl.load_workbook(C.DataDir+"/"+Iws.cell(row = iii, column = 1).value)
			except: return
			Dws = Dwb.active
			checkDate = min(jj,ws.max_row)
			Date = Dws.cell(row = checkDate, column = 1).value

			#print(Date,jj,C.DataDir+"/"+Iws.cell(row = iii, column = 1).value)
			while ((LW -Date).days >= 0):
				checkDate -= 1
				Date = Dws.cell(row = checkDate, column = 1).value
			checkDate -= 1
			if checkDate == 1: return
			MondayValue = Dws.cell(row = checkDate, column = 2).value

			MAXval = min(5000, money[0])

			ant = math.floor( MAXval / (1.0025*MondayValue))
			if ant==0: return
			#print(ant,MondayValue,type(ant))
			temp = ws.cell(row = R, column = 3).value
			money[0] -= ant * MondayValue * 1.0025
			money[1] += ant * MondayValue
			if temp == 1:
				S = [ws.cell(row = R, column = 1).value, 1,ant,MondayValue,Dws.cell(row = checkDate, column = 1).value]
			if temp == 2:
				S = [ws.cell(row = R, column = 1).value, 2,ant,MondayValue,Dws.cell(row = checkDate, column = 1).value]
			if temp == 3:
				S = [ws.cell(row = R, column = 1).value, 1,ant,MondayValue,Dws.cell(row = checkDate, column = 1).value]
				stocks.append(S)
				S = [ws.cell(row = R, column = 1).value, 2,ant,MondayValue,Dws.cell(row = checkDate, column = 1).value]
			#print("Buy "+str(ant) +" "+str(S[0])+" for "+str(MondayValue)+" Tot: "+str(MondayValue*ant)) 
			stocks.append(S)


def indexMaker():
	for N in glob.glob(C.DataDir+"/*.xlsx"):
		M = N.split('/')[-1]
		
		Nwb = openpyxl.load_workbook(N)
		Nws = Nwb.active

		try:
			wb = openpyxl.load_workbook(C.DataDir+"/AAA_index.xlsx")
		except:
			wb = openpyxl.Workbook()

		ws = wb.active

		ws.cell(row = ws.max_row + 1, column = 1).value = M
		ws.cell(row = ws.max_row, column = 2).value = Nws.title

		wb.save(C.DataDir+"/AAA_index.xlsx")

def plotter():
	Mwb = openpyxl.load_workbook(C.DataDir+'/Reports/monneyReport.xlsx')
	Mws = Mwb.active
	R = []
	for L in range(1,Mws.max_row):
		R.append(Mws.cell(row = L, column = 1).value+Mws.cell(row = L, column = 2).value)
	X=range(1,len(R))
	plt.plot(R)
	plt.show()
#indexMaker()
historyTrader()
plotter()
